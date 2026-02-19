from datetime import datetime
import tkinter as tk
from PIL import Image, ImageTk
from num2words import num2words  # Import the num2words library
from tkinter import messagebox
from tkinter import ttk  # Import ttk for Combobox
import openpyxl
from openpyxl import Workbook
import sys
import os
import win32print
import subprocess
from win32con import DMORIENT_LANDSCAPE, DMORIENT_PORTRAIT

# Load and prepare the image
FONT = "Arial"
FSIZE = 12


PLACEHOLDER_DATE = "Enter Date"
PLACEHOLDER_PAYEE  = "Select or Enter Payee"
PLACEHOLDER_AMOUNT = "1234.56"
INVALID_WORDS = ("Invalid input", "One thousand, two hundred and thirty-four")


# Source for the image for APP
def get_resource_path(filename):
    if hasattr(sys, '_MEIPASS'):
        return os.path.join(sys._MEIPASS, filename)
    return os.path.join(os.path.abspath("."), filename)

IMAGE_PATH = get_resource_path("cheq_BG.png")

class ChequeWriterApp:
    def __init__(self, root):
        
        self.root = root
        self.root.title("Cheque Writer")

        # Demo payees for testing
        #self.saved_payees = ["John Doe", "Jane Smith", "Acme Corp"]

        self.categorized_payees = {
            "Supplier": [],
            "Employee": [],
            "Client": [],
            "Other": []
        }

        self.bg_image = Image.open(IMAGE_PATH).convert("RGB")
        self.bg_image = self.bg_image.resize((1000, int(self.bg_image.height * 1000 / self.bg_image.width)))

        self.bg_photo = ImageTk.PhotoImage(self.bg_image)


        img_width, img_height = self.bg_image.size
        print("Image size:", img_width, img_height)

        # Create a Canvas to hold the image and widgets
        self.canvas = tk.Canvas(root, width=self.bg_image.width, height=self.bg_image.height)
        self.canvas.pack()

        # Display the image on the canvas

        self.canvas.create_image(0, 0, image=self.bg_photo, anchor='nw')

        
        ####################################################################################################################################
                                                        # SIDE BAR For Cheq #1 (TOP) #
        ####################################################################################################################################
        # Date
        self.date_label = tk.Label(root, text="", font=(FONT, FSIZE), bg="#E0E0E0", fg="black", bd=1, relief="solid", anchor="w", width=10)
        self.date_label_window = self.canvas.create_window(70, 40, anchor="nw", window=self.date_label)

        # Name
        self.name_label1 = tk.Label(root, text="", font=(FONT, FSIZE), bg="#FFB66C", fg="black", bd=1, relief="solid", anchor="w", width=15)
        self.name_label1_window = self.canvas.create_window(50, 65, anchor="nw", window=self.name_label1)
        # last Name
        self.Lname_label1 = tk.Label(root, text="", font=(FONT, FSIZE), bg="#FFB66C", fg="black", bd=1, relief="solid", anchor="w", width=20)
        self.Lname_label1_window = self.canvas.create_window(50, 85, anchor="nw", window=self.Lname_label1)

        # Amount: Dollar
        self.dollar_label1 = tk.Label(root, text="", font=(FONT, FSIZE), bg="#FFF8DC", fg="black", bd=1, relief="solid", anchor="e", width=7, justify= "right")
        self.dollar_label1_window = self.canvas.create_window(235, 65, anchor="nw", window=self.dollar_label1)
        # Amount: Cents
        self.cent_label1 = tk.Label(root, text="", font=(FONT, FSIZE), bg="#A8D5BA", fg="black", bd=1, relief="solid", anchor="w", width=2)
        self.cent_label1_window = self.canvas.create_window(307, 65, anchor="nw", window=self.cent_label1)

        # FOR
        # Line: 1
        self.for_label1 = tk.Label(root, text="", font=(FONT, FSIZE), bg="#A8D5BA", fg="black", bd=1, relief="solid", anchor="w", width=18)
        self.for_label1_window = self.canvas.create_window(66, 110, anchor="nw",  window=self.for_label1)
        # Line: 2
        self.for_label1_2 = tk.Label(root, text="", font=(FONT, FSIZE), bg="#A8D5BA", fg="black", bd=1, relief="solid", anchor="w", width=22)
        self.for_label1_window_2 = self.canvas.create_window(30, 130, anchor="nw", window=self.for_label1_2)


        ####################################################################################################################################
                                                        # SIDE BAR For Cheq #2 (BOTTOM) #
        ####################################################################################################################################
        # Date
        self.date_label2 = tk.Label(root, text="", font=(FONT, FSIZE), bg="#E0E0E0", fg="black", bd=1, relief="solid", anchor="w", width=10)
        self.date_label2_window = self.canvas.create_window(70, 343, anchor="nw", window=self.date_label2)

        # Name
        self.name_label2 = tk.Label(root, text="", font=(FONT, FSIZE), bg="#FFB66C", fg="black", bd=1, relief="solid", anchor="w", width=15)
        self.name_label2_window = self.canvas.create_window(50, 368, anchor="nw", window=self.name_label2)
        # last Name
        self.Lname_label2 = tk.Label(root, text="", font=(FONT, FSIZE), bg="#FFB66C", fg="black", bd=1, relief="solid", anchor="w", width=20)
        self.Lname_label2_window = self.canvas.create_window(50, 388, anchor="nw", window=self.Lname_label2)

        # Amount: Dollar
        self.dollar_label2 = tk.Label(root, text="", font=(FONT, FSIZE), bg="#FFF8DC", fg="black", bd=1, relief="solid", anchor="e", width=7, justify= "right")
        self.dollar_label2_window2 = self.canvas.create_window(235, 365, anchor="nw", window=self.dollar_label2)
        # Amount: Cents
        self.cent_label2 = tk.Label(root, text="", font=(FONT, FSIZE), bg="#A8D5BA", fg="black", bd=1, relief="solid", anchor="w", width=2)
        self.cent_label2_window = self.canvas.create_window(307, 365, anchor="nw", window=self.cent_label2)

        # FOR
        # Line: 1
        self.for_label2 = tk.Label(root, text="", font=(FONT, FSIZE), bg="#A8D5BA", fg="black", bd=1, relief="solid", anchor="w", width=18)
        self.for_label2_window = self.canvas.create_window(66, 410, anchor="nw", window=self.for_label2)
        # Line: 2
        self.for_label2_2 = tk.Label(root, text="", font=(FONT, FSIZE), bg="#A8D5BA", fg="black", bd=1, relief="solid", anchor="w", width=22)
        self.for_label2_window_2 = self.canvas.create_window(30, 430, anchor="nw", window=self.for_label2_2)

        self.bg = tk.Label(root, text="", font=(FONT, FSIZE), bg = "#EEF5EB", anchor="w", width=200, height=20)
        self.bg = self.canvas.create_window(0, 600, anchor="nw", window=self.bg)
        

        ####################################################################################################################################
                                                        # Frame For Cheq #1 (TOP) #
        ####################################################################################################################################
        # Place Entry box for "DATE"
        self.entry_date1 = tk.Entry(root, font=(FONT, FSIZE), width=10, bd=1)
        self.entry_window = self.canvas.create_window(880, 50, anchor='nw', window=self.entry_date1)
        self.text_preview = self.canvas.create_text(880, 50, anchor='nw', text="", font=(FONT, FSIZE), fill="black")    

        # Add placeholder functionality for entry_date1
        self.entry_date1.insert(0, PLACEHOLDER_DATE )
        self.entry_date1.config(fg="grey")  # Keep the text grey
        self.entry_date1.bind("<FocusIn>", lambda event: self.clear_placeholder(event, self.entry_date1, PLACEHOLDER_DATE ))
        self.entry_date1.bind("<FocusOut>", lambda event: [self.add_placeholder(event, self.entry_date1, PLACEHOLDER_DATE ), self.validate_date(self.entry_date1)])

        self.entry_date1.bind("<KeyRelease>", self.update_date_label)

        # Add the checkbox for today's date
        self.date_checkbox1_var = tk.IntVar()
        self.date_checkbox = tk.Checkbutton(
            root, 
            text="Today", 
            variable=self.date_checkbox1_var, 
            command=self.fill_date_with_today
        )
        self.date_checkbox_window = self.canvas.create_window(820, 50, anchor='nw', window=self.date_checkbox)

        ########################################################################
        
        # Payee Dropdown for Cheque #1
        self.payee_var1 = tk.StringVar(value=PLACEHOLDER_PAYEE)
        self.payee_button1 = tk.Button(
            root,
            textvariable=self.payee_var1,
            font=(FONT, FSIZE),
            width=37,
            command=lambda: self.create_scrollable_dropdown(
                self.payee_button1, self.payee_var1, 415 + self.root.winfo_x(), 93 + self.root.winfo_y()
            )
        )
        self.payee_window1 = self.canvas.create_window(415, 93, anchor='nw', window=self.payee_button1)

        # Manage Button for Payee Dropdown #1
        self.manage_button1 = tk.Button(
            root, 
            text="Manage", 
            font=(FONT, FSIZE), 
            command=self.open_manage_payees_window
        )
        self.manage_button_window1 = self.canvas.create_window(770, 93, anchor='nw', window=self.manage_button1)
        
        ################################################################################
        # Place Entry box for "$____" 
        self.entry_dollar1 = tk.Entry(root, font=(FONT, FSIZE), width=10, bd=1)
        self.entry_window = self.canvas.create_window(880, 110, anchor='nw', window=self.entry_dollar1)
        self.text_preview = self.canvas.create_text(880, 110, anchor='nw', text="", font=(FONT, FSIZE), fill="black")

        # Add placeholder functionality for $____
        self.entry_dollar1.insert(0, PLACEHOLDER_AMOUNT )
        self.entry_dollar1.config(fg="grey")  # Keep the text grey
        self.entry_dollar1.bind("<FocusIn>", lambda event: self.clear_placeholder(event, self.entry_dollar1, PLACEHOLDER_AMOUNT ))
        self.entry_dollar1.bind("<FocusOut>", lambda event: self.add_placeholder(event, self.entry_dollar1, PLACEHOLDER_AMOUNT ))

        self.entry_dollar1.bind(
            "<FocusOut>", 
            lambda event: self.update_amount_in_words(
                event, 
                dollar_entry=self.entry_dollar1, 
                amount_entry=self.entry_amount1, 
                cents_entry=self.entry_100_1
            )
        )
        self.entry_dollar1.bind(
            "<KeyRelease>", 
            lambda event: [self.update_amount_in_words(
                event, 
                dollar_entry=self.entry_dollar1, 
                amount_entry=self.entry_amount1, 
                cents_entry=self.entry_100_1),
                self.update_dollar_label(),
                self.update_cent_label()
                ]
            )

        ################################################################################
        # Place Entry box for "Amount in words" 
        self.entry_amount1 = tk.Entry(root, font=(FONT, FSIZE), width=45, bd=1, state="disabled")
        self.entry_window = self.canvas.create_window(400, 130, anchor='nw', window=self.entry_amount1)
        self.entry_amount1.config(state="normal")  # Temporarily enable to set default value
        self.entry_amount1.insert(0, "One thousand, two hundred and thirty-four")
        self.entry_amount1.config(state="disabled")  # Disable again

        ################################################################################
        # Place Entry box for "___/100" 
        self.entry_100_1 = tk.Entry(root, font=(FONT, FSIZE), width=3, bd=1, state="disabled")
        self.entry_window = self.canvas.create_window(860, 135, anchor='nw', window=self.entry_100_1)
        self.entry_100_1.config(state="normal")  # Temporarily enable to set default value
        self.entry_100_1.insert(0, "56")
        self.entry_100_1.config(state="disabled")  # Disable again

        ################################################################################
        # Place Entry box for "FOR_LINE #1" 
        self.entry_for1_line1 = tk.Entry(root, font=(FONT, FSIZE), width=30, bd=1)
        self.entry_window = self.canvas.create_window(400, 205, anchor='nw', window=self.entry_for1_line1)
        self.text_preview = self.canvas.create_text(400, 205, anchor='nw', text="", font=(FONT, FSIZE), fill="black")

        # Add placeholder functionality for FOR_LINE #1
        self.entry_for1_line1.insert(0, "January 1, 2025 to January 16, 2025")
        self.entry_for1_line1.config(fg="grey")  # Keep the text grey
        self.entry_for1_line1.bind("<FocusIn>", lambda event: self.clear_placeholder(event, self.entry_for1_line1, "January 1, 2025 to January 16, 2025"))
        self.entry_for1_line1.bind("<FocusOut>", lambda event: self.add_placeholder(event, self.entry_for1_line1, "January 1, 2025 to January 16, 2025"))
        
        self.entry_for1_line1.bind("<KeyRelease>", lambda event: self.update_for_label())

        # Place Entry box for "FOR_LINE #2" 
        self.entry_for1_line2 = tk.Entry(root, font=(FONT, FSIZE), width=30, bd=1)
        self.entry_window = self.canvas.create_window(400, 225, anchor='nw', window=self.entry_for1_line2)
        self.text_preview = self.canvas.create_text(400, 225, anchor='nw', text="", font=(FONT, FSIZE), fill="black")

        # Add placeholder functionality for FOR_LINE #2
        self.entry_for1_line2.insert(0, "998877, 665544, 332211")
        self.entry_for1_line2.config(fg="grey")  # Keep the text grey
        self.entry_for1_line2.bind("<FocusIn>", lambda event: self.clear_placeholder(event, self.entry_for1_line2, "998877, 665544, 332211"))
        self.entry_for1_line2.bind("<FocusOut>", lambda event: self.add_placeholder(event, self.entry_for1_line2, "998877, 665544, 332211"))

        self.entry_for1_line2.bind("<KeyRelease>", lambda event: self.update_for_label())

        # Clear Button for Cheque #1
        self.clear_button1 = tk.Button(
            root,
            text="Clear",
            font=(FONT, FSIZE),
            command=lambda: self.clear_cheque(1)
        )
        self.clear_button_window1 = self.canvas.create_window(950, 220, anchor='se', window=self.clear_button1)


        ####################################################################################################################################
                                                        # Frame For Cheq #2 (BOTTOM) #
        ####################################################################################################################################
        # Place Entry box for "DATE2"
        self.entry_date2 = tk.Entry(root, font=(FONT, FSIZE), width=10, bd=1)
        self.entry_window = self.canvas.create_window(880, 350, anchor='nw', window=self.entry_date2)
        self.text_preview = self.canvas.create_text(880, 350, anchor='nw', text="", font=(FONT, FSIZE), fill="black")    

        # Add placeholder functionality for entry_date2
        self.entry_date2.insert(0, PLACEHOLDER_DATE )
        self.entry_date2.config(fg="grey")  # Keep the text grey
        self.entry_date2.bind("<FocusIn>", lambda event: self.clear_placeholder(event, self.entry_date2, PLACEHOLDER_DATE ))
        self.entry_date2.bind("<FocusOut>", lambda event: [self.add_placeholder(event, self.entry_date2, PLACEHOLDER_DATE ), self.validate_date(self.entry_date2)])

        self.entry_date2.bind("<KeyRelease>", self.update_date_label)

        # Add the checkbox for today's date for DATE2
        self.date_checkbox2_var = tk.IntVar()
        self.date_checkbox2 = tk.Checkbutton(
            root, 
            text="Today", 
            variable=self.date_checkbox2_var, 
            command=self.fill_date2_with_today
        )
        self.date_checkbox2_window = self.canvas.create_window(820, 350, anchor='nw', window=self.date_checkbox2)
        
        ########################################################################
        
        # Payee Dropdown for Cheque #2
        self.payee_var2 = tk.StringVar(value=PLACEHOLDER_PAYEE)
        self.payee_button2 = tk.Button(
            root,
            textvariable=self.payee_var2,
            font=(FONT, FSIZE),
            width=37,
            command=lambda: self.create_scrollable_dropdown(
                self.payee_button2, self.payee_var2, 415 + self.root.winfo_x(), 393 + self.root.winfo_y()
            )
        )
        self.payee_window2 = self.canvas.create_window(415, 393, anchor='nw', window=self.payee_button2)

        # Manage Button for Payee Dropdown #2
        self.manage_button2 = tk.Button(
            root, 
            text="Manage", 
            font=(FONT, FSIZE), 
            command=self.open_manage_payees_window
        )
        self.manage_button_window2 = self.canvas.create_window(770, 393, anchor='nw', window=self.manage_button2)
        
        ################################################################################
        # Place Entry box for "$____" 
        self.entry_dollar2 = tk.Entry(root, font=(FONT, FSIZE), width=10, bd=1)
        self.entry_window = self.canvas.create_window(880, 410, anchor='nw', window=self.entry_dollar2)
        self.text_preview = self.canvas.create_text(880, 410, anchor='nw', text="", font=(FONT, FSIZE), fill="black")

        # Add placeholder functionality for $____
        self.entry_dollar2.insert(0, PLACEHOLDER_AMOUNT )
        self.entry_dollar2.config(fg="grey")  # Keep the text grey
        self.entry_dollar2.bind("<FocusIn>", lambda event: self.clear_placeholder(event, self.entry_dollar2, PLACEHOLDER_AMOUNT ))
        self.entry_dollar2.bind("<FocusOut>", lambda event: self.add_placeholder(event, self.entry_dollar2, PLACEHOLDER_AMOUNT ))
        self.entry_dollar2.bind(
            "<FocusOut>", 
            lambda event: self.update_amount_in_words(
                event, 
                dollar_entry=self.entry_dollar2, 
                amount_entry=self.entry_amount2, 
                cents_entry=self.entry_100_2
            )
        )
        self.entry_dollar2.bind(
            "<KeyRelease>", 
            lambda event: [self.update_amount_in_words(
                event, 
                dollar_entry=self.entry_dollar2, 
                amount_entry=self.entry_amount2, 
                cents_entry=self.entry_100_2),
                self.update_dollar_label(),
                self.update_cent_label()]
            )

        ################################################################################
        # Place Entry box for "Amount in words" 
        self.entry_amount2 = tk.Entry(root, font=(FONT, FSIZE), width=45, bd=1, state="disabled")
        self.entry_window = self.canvas.create_window(400, 430, anchor='nw', window=self.entry_amount2)
        self.entry_amount2.config(state="normal")  # Temporarily enable to set default value
        self.entry_amount2.insert(0, "One thousand, two hundred and thirty-four")
        self.entry_amount2.config(state="disabled")  # Disable again

        ################################################################################
        # Place Entry box for "___/100" 
        self.entry_100_2 = tk.Entry(root, font=(FONT, FSIZE), width=3, bd=1, state="disabled")
        self.entry_window = self.canvas.create_window(860, 435, anchor='nw', window=self.entry_100_2)
        self.entry_100_2.config(state="normal")  # Temporarily enable to set default value
        self.entry_100_2.insert(0, "56")
        self.entry_100_2.config(state="disabled")  # Disable again

        ################################################################################
        # Place Entry box for "FOR_LINE #1" 
        self.entry_for2_line1 = tk.Entry(root, font=(FONT, FSIZE), width=30, bd=1)
        self.entry_window = self.canvas.create_window(400, 505, anchor='nw', window=self.entry_for2_line1)
        self.text_preview = self.canvas.create_text(400, 505, anchor='nw', text="", font=(FONT, FSIZE), fill="black")

        # Add placeholder functionality for FOR_LINE #1
        self.entry_for2_line1.insert(0, "January 1, 2023 to January 16, 2023")
        self.entry_for2_line1.config(fg="grey")  # Keep the text grey
        self.entry_for2_line1.bind("<FocusIn>", lambda event: self.clear_placeholder(event, self.entry_for2_line1, "January 1, 2023 to January 16, 2023"))
        self.entry_for2_line1.bind("<FocusOut>", lambda event: self.add_placeholder(event, self.entry_for2_line1, "January 1, 2023 to January 16, 2023"))

        self.entry_for2_line1.bind("<KeyRelease>", lambda event: self.update_for_label())

        # Place Entry box for "FOR_LINE #2" 
        self.entry_for2_line2 = tk.Entry(root, font=(FONT, FSIZE), width=30, bd=1)
        self.entry_window = self.canvas.create_window(400, 525, anchor='nw', window=self.entry_for2_line2)
        self.text_preview = self.canvas.create_text(400, 525, anchor='nw', text="", font=(FONT, FSIZE), fill="black")
        
        # Add placeholder functionality for FOR_LINE #2
        self.entry_for2_line2.insert(0, "998877, 665544, 332211")
        self.entry_for2_line2.config(fg="grey")  # Keep the text grey
        self.entry_for2_line2.bind("<FocusIn>", lambda event: self.clear_placeholder(event, self.entry_for2_line2, "998877, 665544, 332211"))
        self.entry_for2_line2.bind("<FocusOut>", lambda event: self.add_placeholder(event, self.entry_for2_line2, "998877, 665544, 332211"))

        self.entry_for2_line2.bind("<KeyRelease>", lambda event: self.update_for_label())

        # Clear Button for Cheque #2
        self.clear_button2 = tk.Button(
            root,
            text="Clear",
            font=(FONT, FSIZE),
            command=lambda: self.clear_cheque(2)
        )
        self.clear_button_window2 = self.canvas.create_window(950, 520, anchor='se', window=self.clear_button2)

        ####################################################################################################################################
        # Live preview text
        self.text_preview = self.canvas.create_text(435, 320, anchor='nw', text="", font=("Arial", 12), fill="black")

        # Load payees from Excel
        self.load_payees_from_excel()

        self.payee_var1.trace("w", lambda *args: [self.validate_payee_selection(self.payee_var1), self.update_name_label()])
        self.payee_var2.trace("w", lambda *args: [self.validate_payee_selection(self.payee_var2), self.update_name_label()])


        ####################################################################################################################################
                                                                # SETUP PANEL #
        ####################################################################################################################################
        # Frame to group the print controls nicely
        setupFont = "Segoe UI", 11 , "bold"
        self.settings_frame = tk.Frame(root, bg="#EEF5EB")  # same soft background
        self.canvas.create_window(520, 615, anchor="nw", window=self.settings_frame)

        # Orientation label
        orientation_text = f"Orientation: {self.get_printer_orientation()}"
        self.orientation_label = tk.Label(self.settings_frame, text=orientation_text, font=setupFont, bg="#EEF5EB")
        self.orientation_label.grid(row=0, column=1, sticky="w", padx=10, pady=2)

        # Printer label
        printer_name = win32print.GetDefaultPrinter()
        printer_label = tk.Label(self.settings_frame, text=f"Using printer: {printer_name}", font=setupFont, bg="#EEF5EB")
        printer_label.grid(row=1, column=1, sticky="w", padx=10)

        # Checkboxes
        self.cheque1_print_var = tk.BooleanVar()
        self.cheque2_print_var = tk.BooleanVar()
        tk.Checkbutton(self.settings_frame, text="Print Cheque 1", variable=self.cheque1_print_var, font=setupFont, bg="#EEF5EB").grid(row=0, column=0, sticky="w", padx=10)
        tk.Checkbutton(self.settings_frame, text="Print Cheque 2", variable=self.cheque2_print_var, font=setupFont, bg="#EEF5EB").grid(row=1, column=0, sticky="w", padx=10)

        # Buttons
        tk.Button(self.settings_frame, text="Print Selected", font=("Segoe UI", 11), relief="ridge", bg="#F8F8F8", command=self.print_selected_cheques).grid(row=2, column=0, padx=10, pady=6)
        tk.Button(self.settings_frame, text="Open Printer Settings", font=("Segoe UI", 11), relief="ridge", bg="#F8F8F8", command=self.open_printer_preferences).grid(row=2, column=1, padx=10, pady=6)

        # Reminder frame
        reminder_frame = tk.Frame(root, bg="#FFFBEA", bd=2, relief="groove")
        self.canvas.create_window(95, 615, anchor="nw", window=reminder_frame)

        reminder_heading = tk.Label(reminder_frame, text="🛈 Printing Tips:",
                                    font=("Segoe UI", 14, "bold"), bg="#FFFBEA", fg="#2A2A2A", anchor="w")
        reminder_heading.pack(padx=10, pady=(8, 0), anchor="w")

        reminder_text = (
            "• Orientation must be set to Landscape.\n"
            "• Cheque 1 must be selected before printing Cheque 2.\n"
            "• Ensure Date, Payee, and Amount are all filled in.\n"
            "• Use 'Open Printer Settings' to change preferences."
        )

        reminder_label = tk.Label(reminder_frame, text=reminder_text,
                                justify="left", font=("Segoe UI", 12),
                                bg="#FFFBEA", fg="#2A2A2A", anchor="w")
        reminder_label.pack(padx=10, pady=(2, 8), anchor="w")

    ####################################################################################################################################
                                                        # UPDATE FUNCTION ZONE #
                                                              # SIDE BAR #
    ####################################################################################################################################
    def update_preview(self, event=None):
        text = self.entry_payto.get()
        self.canvas.itemconfig(self.text_preview, text=text)

    # Add this method to update the date_label
    def update_date_label(self, event=None):
        # Get the content of entry_date1
        date_text1 = self.entry_date1.get()
        date_text2 = self.entry_date2.get()

        # Only update if user has entered a real date (not empty or placeholder)
        # Cheque #1
        if date_text1 and date_text1 != PLACEHOLDER_DATE :
            self.date_label.config(text=date_text1)
        else:
            self.date_label.config(text="")
        # Cheque #2
        if date_text2 and date_text2 != PLACEHOLDER_DATE :
            self.date_label2.config(text=date_text2)
        else:
            self.date_label2.config(text="")

    # Update Name and Last Name labels based on the selected payee
    def update_name_label(self):
        payee1 = self.payee_var1.get().strip()
        payee2 = self.payee_var2.get().strip()

        # Cheque #1
        if not payee1 or payee1.startswith("---") or payee1 == PLACEHOLDER_PAYEE:
            self.name_label1.config(text="")
            self.Lname_label1.config(text="")
        else:
            parts1 = payee1.split()
            first_name1 = parts1[0]
            last_name1 = parts1[1] if len(parts1) > 1 else ""
            self.name_label1.config(text=first_name1)
            self.Lname_label1.config(text=last_name1)

        # Cheque #2
        if not payee2 or payee2.startswith("---") or payee2 == PLACEHOLDER_PAYEE:
            self.name_label2.config(text="")
            self.Lname_label2.config(text="")
        else:
            parts2 = payee2.split()
            first_name2 = parts2[0]
            last_name2 = parts2[1] if len(parts2) > 1 else ""
            self.name_label2.config(text=first_name2)
            self.Lname_label2.config(text=last_name2)

    # Add this method to update the dollar_label
    def update_dollar_label(self, event=None):
        # Get the content of entry_dollar1 and entry_dollar2
        dollar_value1 = self.entry_dollar1.get().strip()
        dollar_value2 = self.entry_dollar2.get().strip()

        # Only update if the input is not empty and not the placeholder value
        if dollar_value1 and dollar_value1 != PLACEHOLDER_AMOUNT :
            try:
                # Split the value into dollars and cents
                dollars1 = dollar_value1.split('.')[0]  # Get the part before the decimal point
                # Update the dollar_label1 with only the dollar amount
                self.dollar_label1.config(text=dollars1)
            except IndexError:
                # If no decimal point is found, use the entire value
                self.dollar_label1.config(text=dollar_value1)

        if dollar_value2 and dollar_value2 != PLACEHOLDER_AMOUNT :
            try:
                # Split the value into dollars and cents
                dollars2 = dollar_value2.split('.')[0]  # Get the part before the decimal point
                # Update the dollar_label2 with only the dollar amount
                self.dollar_label2.config(text=dollars2)
            except IndexError:
                # If no decimal point is found, use the entire value
                self.dollar_label2.config(text=dollar_value2)

    # Add this method to update the cent_label
    def update_cent_label(self, event=None):
        dollar_value1 = self.entry_dollar1.get().strip()
        dollar_value2 = self.entry_dollar2.get().strip()

        # Cheque #1
        if dollar_value1 and dollar_value1 != PLACEHOLDER_AMOUNT :
            try:
                cents1 = dollar_value1.split('.')[1]
                if len(cents1) == 2 and cents1.isdigit():
                    self.cent_label1.config(text=cents1)
                else:
                    self.cent_label1.config(text="--")
            except IndexError:
                self.cent_label1.config(text="--")
        else:
            self.cent_label1.config(text="")

        # Cheque #2
        if dollar_value2 and dollar_value2 != PLACEHOLDER_AMOUNT :
            try:
                cents2 = dollar_value2.split('.')[1]
                if len(cents2) == 2 and cents2.isdigit():
                    self.cent_label2.config(text=cents2)
                else:
                    self.cent_label2.config(text="--")
            except IndexError:
                self.cent_label2.config(text="--")
        else:
            self.cent_label2.config(text="")

    # Add this method to update the for_label
    def update_for_label(self):
        # ===== Cheque #1 =====
        text1_line1 = self.entry_for1_line1.get().strip()
        text1_line2 = self.entry_for1_line2.get().strip()

        if text1_line1 and text1_line1 != "January 1, 2025 to January 16, 2025":
            self.for_label1.config(text=text1_line1)
        else:
            self.for_label1.config(text="")

        if text1_line2 and text1_line2 != "998877, 665544, 332211":
            self.for_label1_2.config(text=text1_line2)
        else:
            self.for_label1_2.config(text="")

        # ===== Cheque #2 =====
        text2_line1 = self.entry_for2_line1.get().strip()
        text2_line2 = self.entry_for2_line2.get().strip()

        if text2_line1 and text2_line1 != "January 1, 2023 to January 16, 2023":
            self.for_label2.config(text=text2_line1)
        else:
            self.for_label2.config(text="")

        if text2_line2 and text2_line2 != "998877, 665544, 332211":
            self.for_label2_2.config(text=text2_line2)
        else:
            self.for_label2_2.config(text="")

    # Define methods for placeholder handling
    def clear_placeholder(self, event, widget, placeholder=None):
        if isinstance(widget, ttk.Combobox):  # For ttk.Combobox
            if widget.get() == placeholder:
                widget.set("")  # Clear the placeholder
        elif isinstance(widget, tk.Entry):  # For tk.Entry
            if widget.get() == placeholder:
                widget.delete(0, tk.END)
                widget.config(fg="black")  # Set text color to black

    def add_placeholder(self, event, widget, placeholder):
        if isinstance(widget, ttk.Combobox):  # For ttk.Combobox
            if not widget.get():  # If the box is empty
                widget.set(placeholder)  # Restore the placeholder
        elif isinstance(widget, tk.Entry):  # For tk.Entry
            if not widget.get():  # If the box is empty
                widget.insert(0, placeholder)
                widget.config(fg="grey")  # Set text color to grey

    # Method to fill the date with today's date
    # This method is called when the checkbox is checked/unchecked
    def fill_date_with_today(self):
        if self.date_checkbox1_var.get():  # If the checkbox is ticked
            today_date = datetime.now().strftime("%d/%m/%Y")  # Use DD/MM/YYYY format
            self.entry_date1.delete(0, tk.END)  # Clear the DATE box
            self.entry_date1.insert(0, today_date)  # Insert today's date
            self.entry_date1.config(fg="black")  # Set text color to black
    
        else:  # If the checkbox is unticked
            self.entry_date1.delete(0, tk.END)  # Clear the DATE box
            self.entry_date1.insert(0, PLACEHOLDER_DATE )  # Restore placeholder
            self.entry_date1.config(fg="grey")  # Set text color to grey

        self.update_date_label()  # Update the date label with the current date

    # Method to fill the date2 with today's date
    def fill_date2_with_today(self):
        if self.date_checkbox2_var.get():  # If the checkbox is ticked
            today_date = datetime.now().strftime("%d/%m/%Y")  # Use DD/MM/YYYY format
            self.entry_date2.delete(0, tk.END)  # Clear the DATE2 box
            self.entry_date2.insert(0, today_date)  # Insert today's date
            self.entry_date2.config(fg="black")  # Set text color to black
        else:  # If the checkbox is unticked
            self.entry_date2.delete(0, tk.END)  # Clear the DATE2 box
            self.entry_date2.insert(0, PLACEHOLDER_DATE )  # Restore placeholder
            self.entry_date2.config(fg="grey")  # Set text color to grey

        self.update_date_label()  # Update the date label with the current date

    # Method to validate the date format
    def validate_date(self, entry):
        date_value = entry.get()
        try:
            # Check if the input matches the format MM/DD/YYYY
            if not date_value.replace("/", "").isdigit() or date_value.count("/") != 2:
                raise ValueError("Invalid format")

            # Split the date into month, day, and year
            day, month, year = map(int, date_value.split("/"))

            # Check if the year is exactly 4 digits
            if len(str(year)) != 4:
                raise ValueError("Year must be 4 digits")
    
            # Check if the year matches the current year
            #current_year = datetime.now().year
            #if year != current_year:
            #    raise ValueError(f"Invalid year. Year must be {current_year}")

            # Check if the month is valid
            if month < 1 or month > 12:
                raise ValueError("Invalid month")

            # Check if the day is valid for the given month
            if day < 1 or day > 31:
                raise ValueError("Invalid day")

            # Handle months with 30 days
            if month in [4, 6, 9, 11] and day > 30:
                raise ValueError("Invalid day for the given month")

            # Handle February (leap year check)
            if month == 2:
                if (year % 4 == 0 and year % 100 != 0) or (year % 400 == 0):  # Leap year
                    if day > 29:
                        raise ValueError("Invalid day for February in a leap year")
                else:  # Non-leap year
                    if day > 28:
                        raise ValueError("Invalid day for February in a non-leap year")

        except ValueError as e:
            # Show an error message and reset the field
            messagebox.showerror("Invalid Date", f"Error: {e}")
            entry.delete(0, tk.END)
            entry.insert(0, PLACEHOLDER_DATE )
            entry.config(fg="grey")

    # Method to convert the dollar amount to words and cents for both entry_dollar1 and entry_dollar2
    def update_amount_in_words(self, event=None, dollar_entry=None, amount_entry=None, cents_entry=None):
    # Get the value from the specified $____ field
        dollar_value = dollar_entry.get().strip()

        # Fallback to default if empty
        if not dollar_value:
            dollar_value = PLACEHOLDER_AMOUNT 

        try:
            # Split into dollars and cents
            dollars_str, cents_str = dollar_value.split('.')

            # Check if cents length is exactly 2
            if len(cents_str) != 2 or not cents_str.isdigit():
                raise ValueError("Cents must be exactly 2 digits.")

            # Convert to integers
            dollars, cents = int(dollars_str), int(cents_str)

            # Convert dollars to words
            dollars_in_words = num2words(dollars, lang='en').capitalize()

            # Temporarily enable the fields for updates
            amount_entry.config(state="normal")
            cents_entry.config(state="normal")

            # Update amount in words
            amount_entry.delete(0, tk.END)
            amount_entry.insert(0, f"{dollars_in_words}")

            # Update cents
            cents_entry.delete(0, tk.END)
            cents_entry.insert(0, f"{cents:02d}")

            # Disable after update
            amount_entry.config(state="disabled")
            cents_entry.config(state="disabled")

        except (ValueError, IndexError):
            # Handle empty, bad format, or bad cents
            amount_entry.config(state="normal")
            cents_entry.config(state="normal")
            amount_entry.delete(0, tk.END)
            cents_entry.delete(0, tk.END)
            amount_entry.insert(0, "Invalid input")
            cents_entry.insert(0, "--")
            amount_entry.config(state="disabled")
            cents_entry.config(state="disabled")

    ########################################################################
    # Method to Open the Manage Payees Window
    def open_manage_payees_window(self):
        # Check if the "Manage Payees" window is already open
        if hasattr(self, "manage_window") and self.manage_window.winfo_exists():
            self.manage_window.lift()  # Bring the existing window to the front
            return

        # Create the "Manage Payees" window
        self.manage_window = tk.Toplevel(self.root)
        self.manage_window.title("Manage Payees")

        # Set the dimensions of the window
        window_width = 575
        window_height = 450

        # Get the screen width and height
        screen_width = self.root.winfo_screenwidth()
        screen_height = self.root.winfo_screenheight()

        # Calculate the position to center the window
        x = (screen_width // 2) - (window_width // 2)
        y = (screen_height // 2) - (window_height // 2)

        # Set the geometry of the window
        self.manage_window.geometry(f"{window_width}x{window_height}+{x}+{y}")

        # Make the window modal
        self.manage_window.grab_set()  # Prevent interaction with the main window
        self.manage_window.transient(self.root)  # Set the parent window

        # Title Label
        tk.Label(self.manage_window, text="Manage Payees", font=(FONT, FSIZE, "bold")).pack(pady=(10, 5))

        # --- Add New Payee Section ---
        tk.Label(self.manage_window, text="ADD Payees", font=(FONT, FSIZE, "bold")).pack(pady=(10, 5))

        form_frame1 = tk.Frame(self.manage_window)
        form_frame1.pack(pady=5)

        tk.Label(form_frame1, text="First Name:", font=(FONT, FSIZE)).grid(row=0, column=0, padx=5, pady=5, sticky='e')
        first_name_entry = tk.Entry(form_frame1, font=(FONT, FSIZE), width=15)
        first_name_entry.grid(row=0, column=1, padx=5)

        tk.Label(form_frame1, text="Last Name:", font=(FONT, FSIZE)).grid(row=0, column=2, padx=5, pady=5, sticky='e')
        last_name_entry = tk.Entry(form_frame1, font=(FONT, FSIZE), width=15)
        last_name_entry.grid(row=0, column=3, padx=5)

        # --- SAVED Payee Section ---
        form_frame2 = tk.Frame(self.manage_window)
        form_frame2.pack(pady=5)

        tk.Label(form_frame2, text="Category:", font=(FONT, FSIZE)).grid(row=0, column=2, padx=5, sticky='e')
        category_var = tk.StringVar(value="Supplier")

        tk.Label(self.manage_window, text="Saved Payees", font=(FONT, FSIZE, "bold")).pack(pady=(10, 5))

        category_menu = tk.OptionMenu(form_frame2, category_var, "Supplier", "Employee", "Client", "Other")
        category_menu.config(font=(FONT, FSIZE))
        category_menu.grid(row=0, column=3, padx=5)

        def add_payee():
            first = first_name_entry.get().strip()
            last = last_name_entry.get().strip()
            category = category_var.get().strip()

            # Determine the full name based on the category
            if category == "Supplier":
                full_name = first  # Only the first name is required for suppliers
            else:
                full_name = f"{first} {last}".strip()  # Both first and last names are required for other categories

            # Validate input
            if (category == "Supplier" and first) or (category != "Supplier" and first and last):
                # Initialize category list if not present
                if category not in self.categorized_payees:
                    self.categorized_payees[category] = []

                # Check if the payee already exists
                if full_name in self.categorized_payees[category]:
                    messagebox.showinfo("Duplicate Payee", f"The payee '{full_name}' already exists in the '{category}' category.")
                else:
                    # Add the payee to the list
                    self.categorized_payees[category].append(full_name)
                    self.update_payee_dropdowns()
                    self.refresh_payee_display(inner_frame)

                    # Save the new payee to the Excel file
                    try:
                        self.save_payee_to_excel_single(first, last, category)
                    except Exception as e:
                        messagebox.showerror("Error", f"Failed to save to Excel: {e}")

                    # Clear the input fields
                    first_name_entry.delete(0, tk.END)
                    last_name_entry.delete(0, tk.END)
            else:
                if category == "Supplier":
                    messagebox.showerror("Invalid Input", "First Name is required for Suppliers.")
                else:
                    messagebox.showerror("Invalid Input", "Both First Name and Last Name are required.")

        add_button = tk.Button(form_frame1, text="Save", font=(FONT, FSIZE), command=add_payee)
        add_button.grid(row=0, column=6, padx=10)

        # --- Display Saved Payees Section ---
        display_frame = tk.Frame(self.manage_window)
        display_frame.pack(fill=tk.BOTH, expand=True, padx=10, pady=10)

        canvas = tk.Canvas(display_frame)
        scrollbar = tk.Scrollbar(display_frame, orient="vertical", command=canvas.yview)
        inner_frame = tk.Frame(canvas)

        inner_frame.bind(
            "<Configure>",
            lambda e: canvas.configure(scrollregion=canvas.bbox("all"))
        )

        canvas.create_window((0, 0), window=inner_frame, anchor="nw")
        canvas.configure(yscrollcommand=scrollbar.set)

        canvas.pack(side="left", fill="both", expand=True)
        scrollbar.pack(side="right", fill="y")

        # Populate on open
        self.refresh_payee_display(inner_frame)

        # Add Save to Excel Button
        save_excel_button = tk.Button(
            self.manage_window,
            text="Save to Excel",
            font=(FONT, FSIZE),
            command=lambda: self.save_payees_to_excel()
        )
        save_excel_button.pack(pady=10)

    ########################################################################
    # Method to Update Payee Dropdowns
    def update_payee_dropdowns(self):
        # Get the grouped payees
        grouped_payees = self.get_grouped_payees()

        # Update the payee variable for Cheque #1
        if self.payee_var1.get() not in grouped_payees:
            self.payee_var1.set(PLACEHOLDER_PAYEE)

        # Update the payee variable for Cheque #2
        if self.payee_var2.get() not in grouped_payees:
            self.payee_var2.set(PLACEHOLDER_PAYEE)

    def save_payees_to_excel(self, filename="payees.xlsx"):
        try:
            # Create a new workbook or load an existing one
            wb = Workbook()
            ws = wb.active
            ws.title = "Payees"

            # Add headers
            ws.append(["First Name", "Last Name", "Category"])

            # Add payees to the Excel file
            for category, names in self.categorized_payees.items():
                for name in names:
                    if category == "Supplier":
                        first_name, last_name = name, ""  # Only the first name is required for suppliers
                    else:
                        first_name, last_name = name.split(" ", 1) if " " in name else (name, "")
                    ws.append([first_name, last_name, category])

            # Save the workbook
            wb.save(filename)
            messagebox.showinfo("Success", f"Payees saved to {filename}")

        except PermissionError:
            messagebox.showerror("Error", f"Failed to save to {filename}. Please close the file if it is open and try again.")
        except Exception as e:
            messagebox.showerror("Error", f"An unexpected error occurred: {e}")

    def save_payee_to_excel_single(self, first_name, last_name, category, filename="payees.xlsx"):
        # Create a new workbook or load an existing one
        try:
            wb = openpyxl.load_workbook(filename)
            ws = wb.active
        except FileNotFoundError:
            wb = Workbook()
            ws = wb.active
            ws.title = "Payees"
            # Add headers if the file doesn't exist
            ws.append(["First Name", "Last Name", "Category"])

        # Handle cases where the last name is None
        if category == "Supplier":
            last_name = ""  # Leave the last name blank for suppliers

        # Append the new payee
        ws.append([first_name, last_name, category])

        # Save the workbook
        wb.save(filename)

    def delete_payee(self, name, category, inner_frame):
        # Remove the payee from the categorized_payees dictionary
        if category in self.categorized_payees and name in self.categorized_payees[category]:
            self.categorized_payees[category].remove(name)

            # Update the dropdowns and refresh the display
            self.update_payee_dropdowns()
            self.refresh_payee_display(inner_frame)

            # Save the updated list to the Excel file
            self.save_payees_to_excel()

    def load_payees_from_excel(self, filename="payees.xlsx"):
        try:
            # Load the workbook and select the active sheet
            wb = openpyxl.load_workbook(filename)
            ws = wb.active

            # Clear the current categorized_payees dictionary
            self.categorized_payees = {
                "Supplier": [],
                "Employee": [],
                "Client": [],
                "Other": []
            }

            # Read the rows from the Excel file and populate categorized_payees
            for row in ws.iter_rows(min_row=2, values_only=True):  # Skip the header row
                first_name, last_name, category = row
                if category in self.categorized_payees:
                    # Handle None for last name
                    last_name = last_name if last_name is not None else ""
                    full_name = f"{first_name} {last_name}".strip()
                    self.categorized_payees[category].append(full_name)

            # Update the dropdowns
            self.update_payee_dropdowns()

        except FileNotFoundError:
            # If the file doesn't exist, initialize an empty categorized_payees dictionary
            self.categorized_payees = {
                "Supplier": [],
                "Employee": [],
                "Client": [],
                "Other": []
            }
        except Exception as e:
            messagebox.showerror("Error", f"Failed to load payees from Excel: {e}")

    def refresh_payee_display(self, inner_frame):
        for widget in inner_frame.winfo_children():
            widget.destroy()

        for category, names in self.categorized_payees.items():
            if names:
                # Display category label
                cat_label = tk.Label(inner_frame, text=f"{category}:", font=(FONT, FSIZE, "bold"),
                                     bg="#e0e0e0", anchor="w", width=50)
                cat_label.pack(fill="x", pady=(5, 0))

                for name in names:
                    # Frame for payee and delete button
                    payee_frame = tk.Frame(inner_frame)
                    payee_frame.pack(fill="x", pady=2)

                    # Payee name label
                    tk.Label(payee_frame, text=name, font=(FONT, FSIZE), anchor="w", width=40).pack(side="left", padx=5)

                    # Delete button
                    delete_button = tk.Button(
                        payee_frame,
                        text="Delete",
                        font=(FONT, FSIZE),
                        command=lambda n=name, c=category: self.delete_payee(n, c, inner_frame)
                    )
                    delete_button.pack(side="right", padx=5)

    def get_grouped_payees(self):
        grouped_payees = []
        for category, names in self.categorized_payees.items():
            if names:
                grouped_payees.append(f"------------------------ {category} ------------------------")  # Add category header
                grouped_payees.extend(names)  # Add payees under the category
        return grouped_payees

    def validate_payee_selection(self, var):
        selected = var.get()
        if selected.startswith("---"):
            messagebox.showinfo("Invalid Selection", "Please select a valid payee.")
            var.set(PLACEHOLDER_PAYEE)
    
    def create_scrollable_dropdown(self, parent_button, payee_var, x, y):
        dropdown = tk.Toplevel(self.root)
        dropdown.wm_overrideredirect(True)
        dropdown.geometry(f"400x200+{x}+{y}")  # Adjust width/height here

        frame = tk.Frame(dropdown)
        frame.pack(fill="both", expand=True)

        scrollbar = tk.Scrollbar(frame)
        scrollbar.pack(side="right", fill="y")

        listbox = tk.Listbox(frame, yscrollcommand=scrollbar.set, font=(FONT, FSIZE), activestyle="none")
        scrollbar.config(command=listbox.yview)
        listbox.pack(fill="both", expand=True)

        # Add items to the Listbox
        for item in self.get_grouped_payees():
            listbox.insert("end", item)

        # Function to handle hover effect
        def on_hover(event):
            widget = event.widget
            index = widget.index("@%s,%s" % (event.x, event.y))  # Get the index of the item under the mouse
            if index >= 0:  # Ensure the index is valid
                widget.selection_clear(0, "end")  # Clear previous selection
                widget.selection_set(index)  # Highlight the hovered item
                widget.activate(index)  # Set the active item

        # Function to handle selection
        def on_select(event):
            selection = listbox.get(listbox.curselection())
            if selection.startswith("---"):  # Ignore category headers
                messagebox.showinfo("Invalid Selection", "Please select a valid payee.")
                return
            payee_var.set(selection)
            dropdown.destroy()

        # Bind hover and selection events
        listbox.bind("<Motion>", on_hover)  # Highlight item on hover
        listbox.bind("<<ListboxSelect>>", on_select)

        # Close the dropdown when it loses focus
        dropdown.focus_force()
        dropdown.bind("<FocusOut>", lambda e: dropdown.destroy())

    # Method to clear the cheque fields
    def clear_cheque(self, cheque_number):
        if cheque_number == 1:
            # Clear fields for Cheque #1
            self.entry_date1.delete(0, tk.END) # Clear the DATE 
            self.entry_date1.insert(0, PLACEHOLDER_DATE )
            self.entry_date1.config(fg="grey")

            self.payee_var1.set(PLACEHOLDER_PAYEE) # Clear the Payee field

            self.entry_dollar1.delete(0, tk.END) # Clear the $____ field
            self.entry_dollar1.insert(0, PLACEHOLDER_AMOUNT )
            self.entry_dollar1.config(fg="grey")

            self.entry_amount1.config(state="normal") # Clear the Amount in Words field
            self.entry_amount1.delete(0, tk.END)
            self.entry_amount1.insert(0, "One thousand, two hundred and thirty-four")
            self.entry_amount1.config(state="disabled")

            self.entry_100_1.config(state="normal") # Clear the cent field
            self.entry_100_1.delete(0, tk.END)
            self.entry_100_1.insert(0, "")
            self.entry_100_1.config(state="disabled")

            self.entry_for1_line1.delete(0, tk.END)  # Clear For Line 1
            self.entry_for1_line1.insert(0, "January 1, 2025 to January 16, 2025")
            self.entry_for1_line1.config(fg="grey")

            self.entry_for1_line2.delete(0, tk.END) # Clear For Line 2
            self.entry_for1_line2.insert(0, "998877, 665544, 332211")
            self.entry_for1_line2.config(fg="grey")

            self.date_checkbox1_var.set(0) # Uncheck the checkbox

            # SIDE BAR Cheq #1
            self.date_label.config(text="")
            self.name_label1.config(text="")
            self.Lname_label1.config(text="")
            self.dollar_label1.config(text="")
            self.cent_label1.config(text="")
            self.for_label1.config(text="")
            self.for_label1_2.config(text="")

        elif cheque_number == 2:
            # Clear fields for Cheque #2
            self.entry_date2.delete(0, tk.END) # Clear the DATE2 
            self.entry_date2.insert(0, PLACEHOLDER_DATE )
            self.entry_date2.config(fg="grey")

            self.payee_var2.set(PLACEHOLDER_PAYEE) # Clear the Payee field

            self.entry_dollar2.delete(0, tk.END) # Clear the $____ field
            self.entry_dollar2.insert(0, PLACEHOLDER_AMOUNT )
            self.entry_dollar2.config(fg="grey")

            self.entry_amount2.config(state="normal") # Clear the Amount in Words field
            self.entry_amount2.delete(0, tk.END)
            self.entry_amount2.insert(0, "One thousand, two hundred and thirty-four")
            self.entry_amount2.config(state="disabled")

            self.entry_100_2.config(state="normal") # Clear the cent field
            self.entry_100_2.delete(0, tk.END)
            self.entry_100_2.insert(0, "")
            self.entry_100_2.config(state="disabled")

            self.entry_for2_line1.delete(0, tk.END) # Clear For Line 1
            self.entry_for2_line1.insert(0, "January 1, 2023 to January 16, 2023")
            self.entry_for2_line1.config(fg="grey")

            self.entry_for2_line2.delete(0, tk.END) # Clear For Line 2
            self.entry_for2_line2.insert(0, "998877, 665544, 332211")
            self.entry_for2_line2.config(fg="grey")

            self.date_checkbox2_var.set(0) # Uncheck the checkbox

            # SIDE BAR Cheq #2
            self.date_label2.config(text="")
            self.name_label2.config(text="")
            self.Lname_label2.config(text="")
            self.dollar_label2.config(text="")
            self.cent_label2.config(text="")
            self.for_label2.config(text="")
            self.for_label2_2.config(text="")

    ####################################################################################################################################
                                                        # PRINTING FUNCTION ZONE #
    ####################################################################################################################################    
    def print_selected_cheques(self):
        selected_cheques = []

        # Check if the printer is set to landscape orientation
        if not self.is_orientation_landscape():
            messagebox.showwarning("Printer Orientation", "Your printer is not set to Landscape orientation. Please fix it in Printer Settings.")
            return

        # Check if cheque 2 is selected without cheque 1
        if self.cheque2_print_var.get() and not self.cheque1_print_var.get():
            messagebox.showwarning("Print Order", "You must select Cheque 1 before printing Cheque 2.")
            return

        cheque_data = {
            1: {
                "selected": self.cheque1_print_var.get(),
                "date": self.entry_date1.get().strip(),
                "payee": self.payee_var1.get().strip(),
                "amount_words": self.entry_amount1.get().strip()
            },
            2: {
                "selected": self.cheque2_print_var.get(),
                "date": self.entry_date2.get().strip(),
                "payee": self.payee_var2.get().strip(),
                "amount_words": self.entry_amount2.get().strip()
            }
        }

        for num in [1, 2]:
            if not getattr(self, f"cheque{num}_print_var").get():
                continue

            date = getattr(self, f"entry_date{num}").get().strip()
            if not date or date == PLACEHOLDER_DATE:
                messagebox.showerror("Missing Date", f"Cheque {num} is selected but the date is not filled in.")
                return

            payee = getattr(self, f"payee_var{num}").get().strip()
            if not payee or payee == PLACEHOLDER_PAYEE or payee.startswith("---"):
                messagebox.showerror("Invalid Payee", f"Cheque {num} has no valid payee selected.")
                return

            amount_words = getattr(self, f"entry_amount{num}").get().strip()
            if amount_words in INVALID_WORDS:
                messagebox.showerror("Invalid Amount", f"Cheque {num} has an invalid amount.")
                return

            selected_cheques.append(num)

        if not selected_cheques:
            messagebox.showinfo("No Selection", "Please select at least one cheque to print.")
            return

        self.send_to_printer(selected_cheques)

    # Method to send cheques to the printer
    def send_to_printer(self, selected_cheques):
        try:
            import win32print
            import win32ui
            import win32con

            # Get the default printer
            printer_name = win32print.GetDefaultPrinter()
            print(f"Sending to printer: {printer_name}")

            # Create device context
            hdc = win32ui.CreateDC()
            hdc.CreatePrinterDC(printer_name)

            # Get printer DPI
            dpi_x = hdc.GetDeviceCaps(win32con.LOGPIXELSX)
            dpi_y = hdc.GetDeviceCaps(win32con.LOGPIXELSY)

            # Start document
            hdc.StartDoc("Cheque Print")
            hdc.StartPage()

            # Set font (Courier New, 16 pt)
            font = win32ui.CreateFont({
                "name": FONT,
                "height": int(-12 * dpi_y / 72),  # Convert point size to logical units
                "weight": 400
            })
            hdc.SelectObject(font)

            # Draw cheques
            for cheque_number in selected_cheques:
                self.print_cheque(hdc, dpi_x, dpi_y, cheque_number)

            # End job
            hdc.EndPage()
            hdc.EndDoc()
            hdc.DeleteDC()

            messagebox.showinfo("Print", "Cheque(s) sent to the printer.")

        except Exception as e:
            messagebox.showerror("Print Error", f"Failed to print:\n{e}")
            print("Print error:", e)

    # Method to print the cheque details at the specified positions
    def print_cheque(self, hdc, dpi_x, dpi_y, cheque_number):
        # Scale the positions to printer coordinates
        scale_factor_x = dpi_x / 96  # Assuming screen DPI is 96
        scale_factor_y = dpi_y / 96
        scale_factor_y_sidebar = (dpi_y / 96) + int(0.85 * dpi_y)

        # Define the positions of the fields on the screen (in pixels)
        if cheque_number == 1:
            date_pos = (880, 50)
            payee_pos = (415, 93)
            amount_pos = (880, 110)
            amount_words_pos = (400, 130)
            cents_pos = (860, 135)
            for_line1_pos = (400, 205)
            for_line2_pos = (400, 225)

            
            # Scale the positions to printer coordinates FOR CHEQ #1
            date_pos = (int(date_pos[0] * scale_factor_x) + int(0.25 * dpi_x),
                        int(date_pos[1] * scale_factor_y) + int(0.1 * dpi_y))
                        
            payee_pos = (int(payee_pos[0] * scale_factor_x) + int(0.25 * dpi_x), 
                        int(payee_pos[1] * scale_factor_y) + int(0.2 * dpi_y))
            
            amount_pos = (int(amount_pos[0] * scale_factor_x) + int(0.25 * dpi_x), 
                        int(amount_pos[1] * scale_factor_y) + int(0.05 * dpi_y))

            amount_words_pos = (int(amount_words_pos[0] * scale_factor_x), 
                                int(amount_words_pos[1] * scale_factor_y) + int(0.19 * dpi_y))

            cents_pos = (int(cents_pos[0] * scale_factor_x) + int(0.35 * dpi_x), 
                        int(cents_pos[1] * scale_factor_y) + int(0.1 * dpi_y))

            for_line1_pos = (int(for_line1_pos[0] * scale_factor_x), 
                            int(for_line1_pos[1] * scale_factor_y) + int(0.1 * dpi_y))
            
            for_line2_pos = (int(for_line2_pos[0] * scale_factor_x), 
                            int(for_line2_pos[1] * scale_factor_y) + int(0.1 * dpi_y))
            
            # Left-side labels (manually matched to canvas layout)
            left_date_pos     = (int(70 * scale_factor_x),                      int(40 * scale_factor_y) + int(0.1 * dpi_y))
            name1_pos         = (int(50 * scale_factor_x)  + int(0.2 * dpi_x),  int(65 * scale_factor_y) + int(0.1 * dpi_y))
            lname1_pos        = (int(50 * scale_factor_x)  + int(0.2 * dpi_x),  int(85 * scale_factor_y) + int(0.1 * dpi_y))
            dollar1_pos       = (int(235 * scale_factor_x) + int(0.3 * dpi_x),  int(65 * scale_factor_y) + int(1.2 * dpi_y))
            cent1_pos         = (int(307 * scale_factor_x) + int(0.2 * dpi_x),  int(65 * scale_factor_y) + int(1.2 * dpi_y))
            for1_line1_pos    = (int(63 * scale_factor_x),                      int(110 * scale_factor_y) + int(0.1 * dpi_y))
            for1_line2_pos    = (int(30 * scale_factor_x),                      int(130 * scale_factor_y) + int(0.1 * dpi_y))
            
        elif cheque_number == 2:
            date_pos = (880, 350)
            payee_pos = (415, 393)
            amount_pos = (880, 410)
            amount_words_pos = (400, 430)
            cents_pos = (860, 435)
            for_line1_pos = (400, 505)
            for_line2_pos = (400, 525)

            
            # Scale the positions to printer coordinates FOR CHEQ #2
            date_pos2 = (int(date_pos[0] * scale_factor_x) + int(0.25 * dpi_x),
                        int(date_pos[1] * scale_factor_y) + int(0.20 * dpi_y))
                        

            payee_pos2 = (int(payee_pos[0] * scale_factor_x) + int(0.25 * dpi_x), 
                        int(payee_pos[1] * scale_factor_y) + int(0.3 * dpi_y))
            
            amount_pos2 = (int(amount_pos[0] * scale_factor_x) + int(0.25 * dpi_x), 
                        int(amount_pos[1] * scale_factor_y) + int(0.2 * dpi_y))

            amount_words_pos2 = (int(amount_words_pos[0] * scale_factor_x), 
                                int(amount_words_pos[1] * scale_factor_y) + int(0.3 * dpi_y))

            cents_pos2 = (int(cents_pos[0] * scale_factor_x) + int(0.35 * dpi_x), 
                        int(cents_pos[1] * scale_factor_y) + int(0.2 * dpi_y))

            for_line1_pos2 = (int(for_line1_pos[0] * scale_factor_x), 
                            int(for_line1_pos[1] * scale_factor_y) + int(0.23 * dpi_y))
            
            for_line2_pos2 = (int(for_line2_pos[0] * scale_factor_x), 
                            int(for_line2_pos[1] * scale_factor_y) + int(0.2  * dpi_y)) 

            # Left-side labels for cheque 2
            left_date2_pos     = (int(70 * scale_factor_x),                     int(343 * scale_factor_y) + int(0.1 * dpi_y))
            name2_pos         = (int(50 * scale_factor_x)  + int(0.2 * dpi_x),  int(368 * scale_factor_y) + int(0.1 * dpi_y))
            lname2_pos        = (int(50 * scale_factor_x)  + int(0.2 * dpi_x),  int(388 * scale_factor_y) + int(0.1 * dpi_y))
            dollar2_pos       = (int(235 * scale_factor_x) + int(0.3 * dpi_x),  int(365 * scale_factor_y) + int(1.32 * dpi_y))
            cent2_pos         = (int(307 * scale_factor_x) + int(0.2 * dpi_x),  int(365 * scale_factor_y) + int(1.32 * dpi_y))
            for2_line1_pos    = (int(63 * scale_factor_x),                      int(410 * scale_factor_y) + int(0.2  * dpi_y))
            for2_line2_pos    = (int(30 * scale_factor_x),                      int(430 * scale_factor_y) + int(0.2  * dpi_y))   


        # Print the cheque details at the scaled positions
        if cheque_number == 1:
            ############## On Cheque ##############
            hdc.TextOut(date_pos[0], date_pos[1], self.entry_date1.get()) # Date
            hdc.TextOut(payee_pos[0], payee_pos[1], self.payee_var1.get()) # Payee
            hdc.TextOut(amount_pos[0], amount_pos[1], self.entry_dollar1.get()) # Amount in dollars
            hdc.TextOut(amount_words_pos[0], amount_words_pos[1], self.entry_amount1.get()) # Amount in words
            hdc.TextOut(cents_pos[0], cents_pos[1], self.entry_100_1.get()) # Amount in cents

            # FOR line 1 and 2
            for_text1 = self.entry_for1_line1.get().strip()
            for_text2 = self.entry_for1_line2.get().strip()
            # Only print line 1 if it's not placeholder
            if for_text1 and for_text1 != "January 1, 2025 to January 16, 2025":
                hdc.TextOut(for_line1_pos[0], for_line1_pos[1], for_text1)

            # Only print line 2 if it's not placeholder
            if for_text2 and for_text2 != "998877, 665544, 332211":
                hdc.TextOut(for_line2_pos[0], for_line2_pos[1], for_text2)

            ############## Label ##############
            hdc.TextOut(left_date_pos[0], left_date_pos[1], self.date_label.cget("text"))
            hdc.TextOut(name1_pos[0], name1_pos[1], self.name_label1.cget("text"))
            hdc.TextOut(lname1_pos[0], lname1_pos[1], self.Lname_label1.cget("text"))
            hdc.TextOut(dollar1_pos[0], dollar1_pos[1], self.dollar_label1.cget("text"))
            hdc.TextOut(cent1_pos[0], cent1_pos[1], self.cent_label1.cget("text"))
            hdc.TextOut(for1_line1_pos[0], for1_line1_pos[1], self.for_label1.cget("text"))
            hdc.TextOut(for1_line2_pos[0], for1_line2_pos[1], self.for_label1_2.cget("text"))

        elif cheque_number == 2:
            hdc.TextOut(date_pos2[0], date_pos2[1], self.entry_date2.get()) # Date
            hdc.TextOut(payee_pos2[0], payee_pos2[1], self.payee_var2.get()) # Payee
            hdc.TextOut(amount_pos2[0], amount_pos2[1], self.entry_dollar2.get()) # Amount in dollars
            hdc.TextOut(amount_words_pos2[0], amount_words_pos2[1], self.entry_amount2.get()) # Amount in words
            hdc.TextOut(cents_pos2[0], cents_pos2[1], self.entry_100_2.get()) # Amount in cents

            # FOR line 1 and 2
            for2_text1 = self.entry_for2_line1.get().strip()
            for2_text2 = self.entry_for2_line2.get().strip()

            if for2_text1 and for2_text1 != "January 1, 2023 to January 16, 2023":
                hdc.TextOut(for_line1_pos2[0], for_line1_pos2[1], for2_text1)

            if for2_text2 and for2_text2 != "998877, 665544, 332211":
                hdc.TextOut(for_line2_pos2[0], for_line2_pos2[1], for2_text2)

            ############## Label ##############
            hdc.TextOut(left_date2_pos[0], left_date2_pos[1], self.date_label2.cget("text"))
            hdc.TextOut(name2_pos[0], name2_pos[1], self.name_label2.cget("text"))
            hdc.TextOut(lname2_pos[0], lname2_pos[1], self.Lname_label2.cget("text"))
            hdc.TextOut(dollar2_pos[0], dollar2_pos[1], self.dollar_label2.cget("text"))
            hdc.TextOut(cent2_pos[0], cent2_pos[1], self.cent_label2.cget("text"))
            hdc.TextOut(for2_line1_pos[0], for2_line1_pos[1], self.for_label2.cget("text"))
            hdc.TextOut(for2_line2_pos[0], for2_line2_pos[1], self.for_label2_2.cget("text"))

    def open_printer_preferences(self):
        try:
            printer_name = win32print.GetDefaultPrinter()

            # Check if printer is actually connected
            hPrinter = win32print.OpenPrinter(printer_name)
            try:
                printer_info = win32print.GetPrinter(hPrinter, 2)
                status = printer_info["Status"]

                if status != 0:
                    raise Exception("Printer is currently offline or unavailable.")

                subprocess.run(
                    f'rundll32.exe printui.dll,PrintUIEntry /e /n "{printer_name}"',
                    check=True,
                    shell=True
                )
            finally:
                win32print.ClosePrinter(hPrinter)

        except Exception as e:
            messagebox.showerror("Printer Error", f"Cannot open printer preferences.\n\n{str(e)}")

        self.orientation_label.config(text=f"Orientation: {self.get_printer_orientation()}")

    def get_printer_orientation(self):
        try:
            printer_name = win32print.GetDefaultPrinter()
            hPrinter = win32print.OpenPrinter(printer_name)
            try:
                devmode = win32print.GetPrinter(hPrinter, 2)["pDevMode"]
                orientation = devmode.Orientation
                if orientation == DMORIENT_LANDSCAPE:
                    return "Landscape"
                elif orientation == DMORIENT_PORTRAIT:
                    return "Portrait"
                else:
                    return "Unknown"
            finally:
                win32print.ClosePrinter(hPrinter)
        except Exception as e:
            return f"Error: {str(e)}"
        
    def is_orientation_landscape(self):
        try:
            printer_name = win32print.GetDefaultPrinter()
            hPrinter = win32print.OpenPrinter(printer_name)
            try:
                devmode = win32print.GetPrinter(hPrinter, 2)["pDevMode"]
                return devmode.Orientation == DMORIENT_LANDSCAPE
            finally:
                win32print.ClosePrinter(hPrinter)
        except:
            return False  # Fail safe


# Run the app
if __name__ == "__main__":
    root = tk.Tk()
    app = ChequeWriterApp(root)
    root.mainloop()

import win32print
print(win32print.GetDefaultPrinter())


